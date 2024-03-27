import openpyxl, pathlib, math, numpy
import matplotlib.pyplot as plt
import PySimpleGUI as sg
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
def GetDataFormExcel(filePath, wbIndex):
    """Funkcja zwraca listę danych z pliku kompatybilnego z programem excel o wyboldowanych etykietach z podanego arkusza

    :param filePath: ścieżka do piliku excel
    :param wbIndex: index arkusza z danymi (indexowanie od 0)
    :type filePath: pathlib.Path or string
    :type wbIndex: int
    :rtype: int or list
    :return: Funkcja zwraca wartość -1 jeźeli plik o podanej ścieżce nie istnieje, w przeciwnym wypadku funkcja zwraca listę w formie:
        [[wsk_1, [[woj_11, [dana_111, dana_11i], [rok_111, rok_11i]], [woj_1j, [dana_111, dana_11k], [rok_111, rok_11k]]]], ..., [wsk_N, [[woj_11, [dana_111, dana_11M], [rok_111, rok_11M]], [woj_1O, [dana_111, dana_11P], [rok_111, rok_11P]]]]]
        gdzie wsk-tekst wskaźnika; woj-nazwa województwa; rok-rok; dana-dana; i, j, k, N, M, O, P ∈ ℕ.
    """
    # Sprawdzenie poprawności ścieżki
    if not filePath.exists():
        print("PATH ERROR: Path does not exist")
        return -1
    # Pobranie zawartości arkuszów z pliku
    wb = openpyxl.load_workbook(filePath)
    # Sprawdzenie poprawności ścieżki
    try:
        temp = wb.worksheets[wbIndex]
    except IndexError:
        print("INDEX ERROR: Worksheet index out of range")
        return -1

    data = []
    plotsIndex = 0
    # Trawers po wskaźnikach
    for pointerIndex in range(2, wb.worksheets[wbIndex].max_column, 9):
        # Sprawdzenie wyboldowania i zawartości komórki (może być ona wyboldowana i nie zawierać tekstu)
        if wb.worksheets[wbIndex][2][pointerIndex].value is not None:
            data.append([wb.worksheets[wbIndex][2][pointerIndex].value, []])
            seriesIndex = 0
            # Trawers po województwach
            for stateIndex in range(5, wb.worksheets[wbIndex].max_row + 1):
                # Sprawdzenie wyboldowania i zawartości komórki (może być ona wyboldowana i nie zawierać tekstu)
                if wb.worksheets[wbIndex][stateIndex][1] is not None:
                    data[plotsIndex][1].append([wb.worksheets[wbIndex][stateIndex][1].value, [], []])
                    # Trawers po latach
                    for yearIndex in range(pointerIndex, pointerIndex + 9):
                        # Sprawdzenie wyboldowania i zawartości komórki (może być ona wyboldowana i nie zawierać tekstu)
                        if wb.worksheets[wbIndex][3][
                            yearIndex] is not None:
                            data[plotsIndex][1][seriesIndex][1].append(
                                0 if wb.worksheets[wbIndex][stateIndex][yearIndex].value == '-' else
                                wb.worksheets[wbIndex][stateIndex][yearIndex].value)
                            data[plotsIndex][1][seriesIndex][2].append(wb.worksheets[wbIndex][3][yearIndex].value)
                    seriesIndex += 1
            plotsIndex += 1

    return [element for element in data if element[1][0][1]]
data = GetDataFormExcel(pathlib.Path.cwd() / "data.xlsx", 1)
def DrawSubPlot(i, cell, data, dataIterator):
    """Funkcja tworzy podwykres

    :param cell: komórka w matplotlib.figure.Figure
    :param data: lista zagnieżdżona z informacją o danych
    :param dataIterator: iterator po wskaźnikach
    """
    style = ['/', 'o', '.', '*']
    color = ["red", "green", "blue", "orange"]
    # plt.sca(cell)
    barWidth = 0.1
    spacing = numpy.arange(len(data[dataIterator][1][0][1]))
    filler = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    iterator = 0
    cell.cla()
    for state in data[dataIterator][1]:
        temp = (state[1] if (i == -1 or len(spacing) - i < 0) else state[1][0:i] + filler[0:len(spacing) - i])
        cell.bar(spacing + barWidth * iterator, temp, barWidth, label=data[dataIterator][1][iterator][0],
                color=color[iterator], hatch=style[iterator], edgecolor="black")
        iterator += 1
    cell.set_xticks(spacing + barWidth * (len(data[dataIterator][1]) / 2 - 0.5), data[dataIterator][1][0][2])
    cell.legend(loc=8)
    title = data[dataIterator][0][:25] + (data[dataIterator][0][25:] and '...')
    cell.set_title(title.title(), va='top', fontweight='bold')
    cell.set_xlabel('okresy [lata]')
    cell.set_ylabel('przedsiębiorstwa ogółem [%]')
def DrawPlots(data, ax, i):
    """Funkcja towrzy wykresy

    :param i: klatka do stworzenia wykresu
    :return: i-tą klatkę wykresu (dla i = -1 zwraca pełny wykres)
    """
    sizeOfData = len(data)

    if sizeOfData < 1:
        print("No plots to draw")
        return
    if sizeOfData == 1:
        # f, ax = plt.subplots(1, 1)
        DrawSubPlot(i, ax, data, 0)
    else:
        height = int(round(0.80 * math.sqrt(sizeOfData)))
        width = int(math.ceil(sizeOfData / height))
        height = max(height, 2)
        width = max(width, 2)

        # f, ax = plt.subplots(height, width)
        iterator = 0
        for row in ax:
            for col in row:
                if iterator >= sizeOfData:
                    break
                DrawSubPlot(i, col, data, iterator)
                iterator += 1
    return f
def GenerateFigure(data):
    if data == -1:
        print("Nie pobrano danych")
        exit()
    if len(data) == 1:
        f, ax = plt.subplots(1, 1)
    elif len(data) > 1:
        height = int(round(0.80 * math.sqrt(len(data))))
        width = int(math.ceil(len(data) / height))
        height = max(height, 2)
        width = max(width, 2)
        f, ax = plt.subplots(height, width)
        for iterator in range(height * width - len(data)):
            f.delaxes(ax[height - 1, width - iterator - 1])
    else:
        exit()
    plt.tight_layout(h_pad=2, w_pad=3)
    plt.subplots_adjust(left=0.1, bottom=0.1)
    f.set_figwidth(15)
    f.set_figheight(10)
    return f, ax
def DrawFigure(canvas, figure, loc=(0, 0)):
    figure_canvas_agg = FigureCanvasTkAgg(figure, canvas)
    figure_canvas_agg.draw()
    figure_canvas_agg.get_tk_widget().pack(side='bottom', fill='both', expand=0)
    return figure_canvas_agg
def NewData(data):
    result = []
    for pointerIndex in range(len(data)):
        result.append([data[pointerIndex][0], []])
        for stateIndex in range(len(data[pointerIndex][1])):
            if window["-WOJ-" + str(stateIndex)].get():
                result[pointerIndex][1].append([data[pointerIndex][1][stateIndex][0], [], []])
                for yearIndex in range(len(data[pointerIndex][1][stateIndex][2])):
                    if window["-ROK-" + str(yearIndex) + "_" + str(pointerIndex)].get():
                        result[pointerIndex][1][stateIndex][1].append(data[pointerIndex][1][stateIndex][1][yearIndex])
                        result[pointerIndex][1][stateIndex][2].append(data[pointerIndex][1][stateIndex][2][yearIndex])
    return [result[i] for i in range(len(result)) if window["-WSK-" + str(i)].get()]

# ------------------------------------------------------------------- Zadanie 1 ----------------------------------------------------------------------------
def GenerateLayout():
    col1 = [
    [sg.Button('Generuj Wykres', font='Helvetica 14')],
    [sg.Frame("Województwa", [[sg.Checkbox(data[0][1][x][0], key="-WOJ-" + str(x)) for x in range(4)]])],
    ]
    temp = []
    col = [[sg.Frame("Wskaźniki i lata", [[sg.Checkbox(data[0][1][x][0], key="-WOJ-" + str(x)) for x in range(4)]], title_color="blue")]]
    for x in range(14):
        temp.append([sg.Checkbox((data[x][0]).title(), key="-WSK-" + str(x), background_color="slategray", expand_x=True)])
        temp.append([sg.Checkbox(data[x][1][0][2][i], key="-ROK-" + str(i) + "_" + str(x)) for i in range(len(data[x][1][0][2]))])
    col1.append([sg.Frame("Wskaźniki i lata", temp)])
    col2=[
        [sg.Canvas(key='-CANVAS-')]
    ]
    return [[sg.Column(col1, element_justification='l', scrollable=True, size=(400, 1080)), sg.Column(col2, element_justification='r', size=(1000, 1080), expand_x=True, expand_y=True)]]

window = sg.Window('Wykres', GenerateLayout(), location = (5,0), resizable=True, finalize=True)
canvas_elem = window['-CANVAS-']
canvas = canvas_elem.TKCanvas

printFlag = False
newData = data
f, ax = GenerateFigure(newData)
delay = 1000
while True:
    event, values = window.read(timeout=delay)
    if event == sg.WINDOW_CLOSED:
        break
    if event == 'Generuj Wykres':
        i = 1
        newData = NewData(data)
        f.clf()
        f, ax = GenerateFigure(newData)
        fig_agg = DrawFigure(canvas, DrawPlots(newData, ax,-1))
        fig_agg.draw()
window.close()

# ------------------------------------------------------------------- Zadanie 2 ----------------------------------------------------------------------------
def GenerateLayout():
    col1 = [
    [sg.Button('Generuj Wykres', font='Helvetica 14')],
    [sg.Text('Milisekundy na klatkę animacji'), sg.Slider(range=(100, 2000), default_value=500, enable_events=True, orientation='horizontal', key='-SLIDER-')],
    # [sg.Text('Województwa')],
    # [sg.Checkbox(data[0][1][x][0], key="-WOJ-" + str(x)) for x in range(4)],
    [sg.Frame("Województwa", [[sg.Checkbox(data[0][1][x][0], key="-WOJ-" + str(x)) for x in range(4)]])],
    ]
    temp = []
    col = [[sg.Frame("Wskaźniki i lata", [[sg.Checkbox(data[0][1][x][0], key="-WOJ-" + str(x)) for x in range(4)]], title_color="blue")]]
    for x in range(14):
        temp.append([sg.Checkbox((data[x][0]).title(), key="-WSK-" + str(x), background_color="slategray", expand_x=True)])
        temp.append([sg.Checkbox(data[x][1][0][2][i], key="-ROK-" + str(i) + "_" + str(x)) for i in range(len(data[x][1][0][2]))])
    col1.append([sg.Frame("Wskaźniki i lata", temp)])
    col2=[
        [sg.Canvas(key='-CANVAS-')]
    ]
    return [[sg.Column(col1, element_justification='l', scrollable=True, size=(400, 1080)), sg.Column(col2, element_justification='r', size=(1000, 1080), expand_x=True, expand_y=True)]]

window = sg.Window('Wykres', GenerateLayout(), location = (5,0), resizable=True, finalize=True)
canvas_elem = window['-CANVAS-']
canvas = canvas_elem.TKCanvas

printFlag = False
i = 0
newData = data
f, ax = GenerateFigure(newData)
delay = 500
while True:
    event, values = window.read(timeout=delay)
    if event == sg.WINDOW_CLOSED:
        break
    if event == 'Generuj Wykres':
        i = 1
        newData = NewData(data)
        f.clf()
        f, ax = GenerateFigure(newData)
        fig_agg = DrawFigure(canvas, DrawPlots(newData, ax,1))
        printFlag = True
    if event == "-SLIDER-":
        delay = int(values['-SLIDER-'])
    elif event == sg.TIMEOUT_EVENT and printFlag:
        f = DrawPlots(newData, ax, i)
        fig_agg.draw()
        i += 1
        if i >= 10:
            i = 1

window.close()