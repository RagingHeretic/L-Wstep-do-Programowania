# Hubert Jackowski

import numpy
import openpyxl, pathlib, math, matplotlib.pyplot as plt
from matplotlib import animation

def GetDataFormExcel(filePath, wbIndex):
    """Funkcja zwraca listę danych z pliku kompatybilnego z programem excel o wyboldowanych etykietach z podanego arkusza
    
    :param filePath: ścieżka do piliku excel
    :param wbIndex: index arkusza z danymi (indexowanie od 0)
    :type filePath: pathlib.Path or string
    :type wbIndex: int
    :rtype: int or list
    :return: Funkcja zwraca wartość -1 jeźeli plik o podanej ścieżce nie istnieje, w przeciwnym wypadku funkcja zwraca listę w formie:
        [[wsk_1, [[woj_11, [dana_111, dana_11i], [rok_111, rok_11i]], [woj_1j, [dana_111, dana_11k], [rok_111, rok_11k]]]], ..., [wsk_N, [[woj_11, [dana_111, dana_11M], [rok_111, rok_11M]], [woj_1O, [dana_111, dana_11P], [rok_111, rok_11P]]]]]
        gdzie wsk-tekst wskaźnika; woj-nazwa województwa; rok-rok; dana-dana; i, j, k, N, M, O, P ∈ ℕ.
    """
    # Sprawdzenie poprawności ścieżki
    if not filePath.exists():
        print("PATH ERROR: Path does not exist")
        return -1
    # Pobranie zawartości arkuszów z pliku
    wb = openpyxl.load_workbook(filePath)
    # Sprawdzenie poprawności ścieżki
    try:
        temp = wb.worksheets[wbIndex]
    except IndexError:
        print("INDEX ERROR: Worksheet index out of range")
        return -1

    data = []
    plotsIndex = 0
    # Trawers po wskaźnikach
    for pointerIndex in range(2, wb.worksheets[wbIndex].max_column, 9):
        # Sprawdzenie wyboldowania i zawartości komórki (może być ona wyboldowana i nie zawierać tekstu)
        if wb.worksheets[wbIndex][2][pointerIndex].font.b and wb.worksheets[wbIndex][2][pointerIndex].value is not None:
            data.append([wb.worksheets[wbIndex][2][pointerIndex].value, []])
            seriesIndex = 0
            # Trawers po województwach
            for stateIndex in range(5, wb.worksheets[wbIndex].max_row+1):
                # Sprawdzenie wyboldowania i zawartości komórki (może być ona wyboldowana i nie zawierać tekstu)
                if wb.worksheets[wbIndex][stateIndex][1].font.b and wb.worksheets[wbIndex][stateIndex][1] is not None:
                    data[plotsIndex][1].append([wb.worksheets[wbIndex][stateIndex][1].value, [], []])
                    # Trawers po latach
                    for yearIndex in range(pointerIndex, pointerIndex + 9):
                        # Sprawdzenie wyboldowania i zawartości komórki (może być ona wyboldowana i nie zawierać tekstu)
                        if wb.worksheets[wbIndex][3][yearIndex].font.b and wb.worksheets[wbIndex][3][yearIndex] is not None:
                            data[plotsIndex][1][seriesIndex][1].append(0 if wb.worksheets[wbIndex][stateIndex][yearIndex].value == '-' else wb.worksheets[wbIndex][stateIndex][yearIndex].value)
                            data[plotsIndex][1][seriesIndex][2].append(wb.worksheets[wbIndex][3][yearIndex].value)
                    seriesIndex += 1
            plotsIndex += 1

    return [element for element in data if element[1][0][1]]

data = GetDataFormExcel(pathlib.Path.cwd() / "data.xlsx", 1)

def DrawSubPlot(i, cell, data, dataIterator):
    """Funkcja tworzy podwykres

    :param cell: komórka w matplotlib.figure.Figure
    :param data: lista zagnieżdżona z informacją o danych
    :param dataIterator: iterator po wskaźnikach
    """
    style = ['/',  'o', '.', '*']
    color = ["red", "green", "blue", "orange"]
    plt.sca(cell)
    barWidth = 0.1
    spacing = numpy.arange(len(data[dataIterator][1][0][1]))
    filler = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    iterator = 0
    plt.cla()
    for state in data[dataIterator][1]:
        temp = (state[1] if (i == -1 or len(spacing)-i < 0) else state[1][0:i] + filler[0:len(spacing)-i])
        plt.bar(spacing + barWidth * iterator, temp, barWidth, label=data[dataIterator][1][iterator][0], color=color[iterator], hatch=style[iterator], edgecolor="black")
        iterator += 1
    plt.xticks(spacing + barWidth * (len(data[dataIterator][1]) / 2 - 0.5), data[dataIterator][1][0][2])
    plt.legend(loc=8)
    title = data[dataIterator][0][:25] + (data[dataIterator][0][25:] and '...')
    plt.title(title.title(), va='top', fontweight='bold')
    plt.xlabel('okresy [lata]')
    plt.ylabel('przedsiębiorstwa ogółem [%]')
def DrawPlots(i):
    """Funkcja towrzy wykresy

    :param i: klatka do stworzenia wykresu
    :return: i-tą klatkę wykresu (dla i = -1 zwraca pełny wykres)
    """
    sizeOfData = len(data)

    if sizeOfData < 1:
        print("No plots to draw")
        return
    if sizeOfData == 1:
        # f, ax = plt.subplots(1, 1)
        DrawSubPlot(i, ax, data, 0)
    else:
        height = int(round(0.80 * math.sqrt(sizeOfData)))
        width = int(math.ceil(sizeOfData / height))
        height = max(height, 2)
        width = max(width, 2)

        # f, ax = plt.subplots(height, width)
        iterator = 0
        for row in ax:
            for col in row:
                if iterator >= sizeOfData:
                    break
                DrawSubPlot(i, col, data, iterator)
                iterator += 1
    return f

if data == -1:
    print("Nie pobrano danych")
    exit()
if len(data) == 1:
    f, ax = plt.subplots(1, 1)
elif len(data) > 1:
    height = int(round(0.80 * math.sqrt(len(data))))
    width = int(math.ceil(len(data) / height))
    height = max(height, 2)
    width = max(width, 2)
    f, ax = plt.subplots(height, width)
    for iterator in range(height * width - len(data)):
        f.delaxes(ax[height - 1, width - iterator - 1])
else:
    exit()
plt.tight_layout()
plt.get_current_fig_manager().window.state('zoomed')

#  ----------------------------------------------- Zadanie 1 -----------------------------------------------
# f = DrawPlots(-1)
# plt.show()
#  ----------------------------------------------- Zadanie 2 -----------------------------------------------
ani = animation.FuncAnimation(fig=f, func=DrawPlots, blit=False, frames=range(1,9), interval=500, repeat=True)
plt.show()
